"""한글화 결과 → Excel 바이트로 변환 (Streamlit st.download_button 용)"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def results_to_xlsx_bytes(results: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "전성분 한글화 결과"

    headers = ["입력", "방향", "한글명", "영문명", "107 기준", "107 상세", "CAS No.", "기원·정의", "이명", "상태"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for ri, r in enumerate(results, 2):
        direction = "영→한" if r.get("direction") == "eng→kor" else "한→영"
        ws.cell(row=ri, column=1, value=r.get("input", ""))
        ws.cell(row=ri, column=2, value=direction)
        ws.cell(row=ri, column=3, value=r.get("kor", ""))
        ws.cell(row=ri, column=4, value=r.get("eng", ""))
        ws.cell(row=ri, column=5, value=r.get("cs_label", ""))
        ws.cell(row=ri, column=6, value=r.get("cs_detail", ""))
        ws.cell(row=ri, column=7, value=r.get("cas", ""))
        ws.cell(row=ri, column=8, value=r.get("origin", ""))
        ws.cell(row=ri, column=9, value=r.get("synonym", ""))
        sc = ws.cell(row=ri, column=10, value=r.get("status", ""))

        cs_status = r.get("cs_status")
        if cs_status == "PROHIBITED":
            fill = PatternFill(start_color="FBE5E5", end_color="FBE5E5", fill_type="solid")
            for col in range(1, 11):
                ws.cell(row=ri, column=col).fill = fill
        elif cs_status in ("RESTRICTED", "IMPURITY"):
            fill = PatternFill(start_color="FFF4D9", end_color="FFF4D9", fill_type="solid")
            for col in range(1, 11):
                ws.cell(row=ri, column=col).fill = fill
        elif r.get("status") == "PARTIAL":
            sc.fill = PatternFill(start_color="E8F1FA", end_color="E8F1FA", fill_type="solid")
        elif r.get("status") == "NOT_FOUND":
            sc.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for c in range(1, 11):
            cell = ws.cell(row=ri, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=10)
        ws.row_dimensions[ri].height = 42

    widths = [18, 8, 20, 24, 14, 32, 12, 36, 18, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "1:1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def default_filename() -> str:
    return f"성분명_사전_{datetime.now():%Y%m%d_%H%M}.xlsx"
